"""
Excel 종목 메모 → stock_notes DB 마이그레이션
실행: python migrate_stock_notes.py
"""
import re
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("openpyxl 필요: pip install openpyxl")
    sys.exit(1)

EXCEL_PATH = Path("/Users/msim/Downloads/메모통합 master.xlsx")
DB_PATH = ".data/news.db"

# 메모 내 날짜 패턴 (M/D, (YY.MM.DD), YYYY/MM/DD, YYYY-MM-DD)
DATE_RE = re.compile(
    r'\((\d{2})\.(\d{2})\.(\d{2})\)'      # (25.12.18) → 2025-12-18
    r'|(\d{4})[/-](\d{2})[/-](\d{2})'     # 2025/11/19 or 2025-11-19
    r'|(\d{1,2})/(\d{1,2})'               # 4/29 or 1/2 (연도 추론 필요)
)


def infer_year(row_date: datetime, month: int, day: int) -> int:
    """엑셀 행의 날짜 기준으로 메모 내 M/D 연도 추론. 미래면 전년도."""
    base = row_date.year
    candidate = datetime(base, month, day)
    # 행 날짜보다 3개월 이상 미래면 전년도
    if (candidate - row_date).days > 90:
        return base - 1
    return base


def clean_memo(text: str) -> str:
    """오탈자 정리: 연속 공백/줄바꿈 정리, 물음표 특수문자 제거 등."""
    # ?로 인코딩된 특수문자(·, …, ' 등)는 원문 의미 보존 — 그대로 둠
    # 앞뒤 공백 제거
    text = text.strip()
    # 3개 이상 연속 줄바꿈 → 2개로 압축
    text = re.sub(r'\n{3,}', '\n\n', text)
    # 줄 끝 trailing space 제거
    text = '\n'.join(line.rstrip() for line in text.split('\n'))
    return text


def split_by_dates(memo: str, row_date: datetime) -> list:
    """
    메모 텍스트를 날짜 태그 기준으로 분리.
    각 항목: {'note_date': 'YYYY-MM-DD', 'content': str}
    날짜 없는 도입부는 note_date = row_date (엑셀 행 날짜)
    """
    # 날짜 토큰 위치 목록 수집
    tokens = []
    for m in DATE_RE.finditer(memo):
        pos = m.start()
        # 패턴 매칭 그룹별 파싱
        if m.group(1):  # (YY.MM.DD)
            yy, mm, dd = int(m.group(1)), int(m.group(2)), int(m.group(3))
            year = 2000 + yy
            try:
                d = datetime(year, mm, dd)
                tokens.append((pos, m.end(), d.strftime('%Y-%m-%d')))
            except ValueError:
                pass
        elif m.group(4):  # YYYY/MM/DD or YYYY-MM-DD
            yy, mm, dd = int(m.group(4)), int(m.group(5)), int(m.group(6))
            try:
                d = datetime(yy, mm, dd)
                tokens.append((pos, m.end(), d.strftime('%Y-%m-%d')))
            except ValueError:
                pass
        elif m.group(7):  # M/D
            month, day = int(m.group(7)), int(m.group(8))
            if 1 <= month <= 12 and 1 <= day <= 31:
                try:
                    year = infer_year(row_date, month, day)
                    d = datetime(year, month, day)
                    tokens.append((pos, m.end(), d.strftime('%Y-%m-%d')))
                except ValueError:
                    pass

    if not tokens:
        # 날짜 없는 메모 전체를 하나로
        content = clean_memo(memo)
        if content:
            return [{'note_date': row_date.strftime('%Y-%m-%d'), 'content': content}]
        return []

    results = []

    # 첫 날짜 토큰 전 텍스트 (도입부)
    first_pos = tokens[0][0]
    intro = memo[:first_pos].strip()
    if intro:
        results.append({
            'note_date': row_date.strftime('%Y-%m-%d'),
            'content': clean_memo(intro),
        })

    # 날짜 토큰 기준 분할
    for i, (start, end, date_str) in enumerate(tokens):
        next_start = tokens[i + 1][0] if i + 1 < len(tokens) else len(memo)
        segment = memo[start:next_start].strip()
        # 날짜 토큰 자체는 content에 포함 (날짜 정보도 텍스트로 보존)
        content = clean_memo(segment)
        if content:
            results.append({'note_date': date_str, 'content': content})

    return results


def run():
    from news_storage import NewsStorage

    print(f"Excel 읽는 중: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Sheet1']
    total_rows = ws.max_row
    print(f"총 {total_rows}개 종목")

    ns = NewsStorage(DB_PATH)

    # stock_master에 종목명 upsert (없는 종목 등록)
    all_rows = list(ws.iter_rows(min_row=1, max_row=total_rows, values_only=True))

    import sqlite3
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    # 기존 notes 초기화 (재실행 시 중복 방지)
    existing = conn.execute("SELECT COUNT(*) FROM stock_notes WHERE source='import'").fetchone()[0]
    if existing > 0:
        print(f"기존 import 노트 {existing}개 발견. 삭제 후 재임포트합니다.")
        conn.execute("DELETE FROM stock_notes WHERE source='import'")
        conn.commit()

    conn.close()

    notes_to_insert = []
    stock_upserts = []

    for row in all_rows:
        code, name, memo, row_date = row
        if not code or not memo:
            continue

        code = str(code).strip().zfill(6)
        name = str(name).strip() if name else ''

        # stock_master upsert (이름만)
        stock_upserts.append((code, name))

        # 날짜 기준 분할
        if not isinstance(row_date, datetime):
            try:
                row_date = datetime.strptime(str(row_date)[:10], '%Y-%m-%d')
            except Exception:
                row_date = datetime.today()

        splits = split_by_dates(str(memo), row_date)
        for s in splits:
            notes_to_insert.append({
                'stock_code': code,
                'note_date': s['note_date'],
                'content': s['content'],
                'source': 'import',
            })

    # stock_master 등록
    with sqlite3.connect(DB_PATH) as conn:
        for code, name in stock_upserts:
            conn.execute(
                """INSERT INTO stock_master (stock_code, stock_name, updated_at)
                   VALUES (?, ?, CURRENT_TIMESTAMP)
                   ON CONFLICT(stock_code) DO UPDATE SET
                       stock_name=COALESCE(excluded.stock_name, stock_master.stock_name),
                       updated_at=CURRENT_TIMESTAMP""",
                (code, name)
            )

    print(f"stock_master {len(stock_upserts)}개 등록/갱신")

    # stock_notes bulk insert
    inserted = ns.bulk_insert_stock_notes(notes_to_insert)
    print(f"stock_notes {inserted}개 노트 입력 (총 {len(notes_to_insert)}개 파싱됨)")

    # 샘플 확인
    sample_code = all_rows[0][0] if all_rows else None
    if sample_code:
        sample_notes = ns.get_stock_notes(str(sample_code).zfill(6))
        print(f"\n샘플 ({sample_code}):")
        for n in sample_notes[:3]:
            print(f"  [{n['note_date']}] {n['content'][:80]}...")


if __name__ == '__main__':
    run()
